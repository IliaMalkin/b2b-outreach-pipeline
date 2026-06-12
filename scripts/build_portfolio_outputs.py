from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_CSV = ROOT / "data" / "outreach_pipeline_masked.csv"
OUT_XLSX = ROOT / "outputs" / "outreach_pipeline_masked.xlsx"


def read_rows() -> list[dict[str, str]]:
    with DATA_CSV.open(encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def build_workbook(rows: list[dict[str, str]]) -> Workbook:
    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    sheet = workbook.active
    sheet.title = "Outreach base"
    headers = list(rows[0])
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        sheet.append([row[column] for column in headers])

    widths = [24, 32, 28, 24, 22, 26, 72, 32, 30, 32, 34]
    for index, width in enumerate(widths[: len(headers)], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    sequence = workbook.create_sheet("Email sequence")
    sequence.append(["Step", "Subject", "Body"])
    for cell in sequence[1]:
        cell.fill = header_fill
        cell.font = header_font
    sequence_rows: list[tuple[str, str, str]] = [
        (
            "Email 1",
            "{{Компания}} and outbound email",
            "{{Имя}}, hello.\n\n{{персонализация}}\n\n"
            "We help B2B teams test outbound email without turning it into a large, risky campaign from day one: "
            "segment, offer, sequence, domain setup, and reply tracking.\n\n"
            "For {{Компания}}, I would start with a small pilot: 2-3 segments, a separate offer for each, "
            "and clear reply metrics before scaling.\n\n"
            "Would it be useful if I showed what that pilot could look like for your market?",
        ),
        (
            "Email 2",
            "Re: {{Компания}} and outbound email",
            "{{Имя}}, quick follow-up.\n\n"
            "When cold email does not work, the problem is often not the channel itself. Usually the list is too broad, "
            "the domain is cold, or the email reads like a generic template.\n\n"
            "The safer version is to test in small batches: validate the segment, keep the copy specific, "
            "track replies, and only then scale.\n\n"
            "If this is relevant, I can share a short outline for a pilot campaign.",
        ),
        (
            "Email 3",
            "Re: {{Компания}} and outbound email",
            "{{Имя}}, last note from me.\n\n"
            "Cold email works best when it is treated as a series of small tests: segment, offer, message, reply quality. "
            "That also keeps the risk low before a bigger rollout.\n\n"
            "If this becomes relevant later, you can reply to this email and I will send over a compact campaign outline.",
        ),
    ]
    for sequence_row in sequence_rows:
        sequence.append(sequence_row)
    sequence.column_dimensions["A"].width = 20
    sequence.column_dimensions["B"].width = 40
    sequence.column_dimensions["C"].width = 96
    for row in sequence.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    qa = workbook.create_sheet("QA summary")
    qa_rows: list[list[object]] = [
        ["Metric", "Value"],
        ["Rows", len(rows)],
        ["Email visibility", "all emails masked in public portfolio"],
        ["Personalization", "filled for every row"],
        ["Contact names", "redacted in public portfolio"],
        ["SMTP validation", "not included in public demo"],
        ["Agent workflow", "documented in docs/agent_workflow.md"],
    ]
    for qa_row in qa_rows:
        qa.append(qa_row)
    for cell in qa[1]:
        cell.fill = header_fill
        cell.font = header_font
    qa.column_dimensions["A"].width = 34
    qa.column_dimensions["B"].width = 70

    return workbook


def main() -> int:
    rows = read_rows()
    if not rows:
        raise RuntimeError("No rows found in masked CSV")
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(rows)
    workbook.save(OUT_XLSX)
    print(f"saved {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
